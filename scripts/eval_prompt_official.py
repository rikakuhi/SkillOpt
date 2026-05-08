#!/usr/bin/env python3
"""Standalone eval: OFFICIAL prompt (SpreadsheetBench original) on verified-400.

Usage:
    python scripts/eval_prompt_official.py --workers 8
    python scripts/eval_prompt_official.py --workers 32 --limit 20
"""
from __future__ import annotations

import argparse
import glob
import json
import os
import random
import re
import subprocess
import sys
import tempfile
import textwrap
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FuturesTimeoutError

import openpyxl

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from reflact.model import (
    chat_messages_with_deployment,
    configure_azure_openai,
    set_backend,
    set_student_deployment,
)
from reflact.envs.spreadsheetbench.evaluator import evaluate


# ── Config ──────────────────────────────────────────────────────────────────

DATA_ROOT = "/home/azureuser/workspace-yqh/sr/spreadsheetbench/data/spreadsheetbench_verified_400"
JSONL_PATH = os.path.join(DATA_ROOT, "dataset.json")
MODEL = "gpt-5-mini"

# ── Official Prompt (from SpreadsheetBench src/prompt.py) ───────────────────

_SYSTEM_PROMPT = (
    "You are an expert Python programmer specializing in spreadsheet manipulation. "
    "You will be given a user instruction together with a preview of an input .xlsx file. "
    "Your job is to write a single self-contained Python script that reads the input file "
    "at the path stored in the variable INPUT_PATH, performs the requested manipulation, "
    "and saves the result to OUTPUT_PATH. Use only the standard library, openpyxl, and pandas. "
    "Do not print anything. Do not use input(). Do not hardcode file paths. "
    "Return ONLY the Python code inside a single ```python ... ``` fenced block."
)


def build_system(skill_content: str = "") -> str:
    base = _SYSTEM_PROMPT
    if skill_content.strip():
        base += f"\n\n## Skill\n{skill_content.strip()}"
    return base


def build_user(instruction, input_xlsx, instruction_type="", answer_position=""):
    try:
        preview = _preview_workbook(input_xlsx)
    except Exception as e:
        preview = f"(failed to preview: {e})"
    extra = ""
    if instruction_type:
        extra += f"\nInstruction type: {instruction_type}"
    if answer_position:
        extra += f"\nExpected answer position: {answer_position}"
    return (
        f"# Instruction\n{instruction}\n{extra}\n\n"
        f"# Input spreadsheet preview\n{preview}\n\n"
        "# Task\n"
        "Write a Python script that reads the workbook from the variable `INPUT_PATH`, "
        "applies the instruction, and writes the modified workbook to `OUTPUT_PATH`. "
        "Preserve all other cells unchanged. "
        "The preview may be truncated — do not hardcode row counts or assume the data ends at the last previewed row; "
        "iterate over all actual rows in the workbook instead. "
        "Return only a ```python``` code block."
    )


# ── Shared utilities (identical to custom version) ──────────────────────────

def _preview_workbook(path, max_rows=5, max_cols=20):
    wb = openpyxl.load_workbook(path, data_only=False)
    chunks = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        chunks.append(f"## Sheet: {sn}  (dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column})")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows),
                                max_col=min(ws.max_column, max_cols), values_only=False):
            cells = []
            for c in row:
                v = c.value
                s = "" if v is None else str(v)
                if len(s) > 40: s = s[:37] + "..."
                cells.append(f"{c.coordinate}={s}")
            chunks.append(" | ".join(cells))
        if ws.max_row > max_rows:
            chunks.append(f"... ({ws.max_row - max_rows} more rows)")
        chunks.append("")
    wb.close()
    return "\n".join(chunks)


def extract_code(text):
    if "```" not in text:
        return text.strip()
    start = text.find("```")
    nl = text.find("\n", start)
    end = text.find("```", nl + 1)
    if nl == -1 or end == -1:
        return text.strip()
    return text[nl + 1:end].strip()


_PATH_RE = re.compile(r'^\s*(INPUT_PATH|OUTPUT_PATH)\s*=\s*.+$', re.MULTILINE)

def strip_paths(code):
    return _PATH_RE.sub("", code)


RUNNER_TEMPLATE = textwrap.dedent("""
    import os, sys, traceback
    INPUT_PATH = {input_path!r}
    OUTPUT_PATH = {output_path!r}
    try:
    {code_indented}
    except Exception:
        traceback.print_exc()
        sys.exit(2)
""")


def run_code(code, input_path, output_path, timeout=120):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    cleaned = strip_paths(code)
    indented = textwrap.indent(cleaned, "    ")
    script = RUNNER_TEMPLATE.format(input_path=input_path, output_path=output_path, code_indented=indented)
    with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False) as f:
        f.write(script)
        tmp = f.name
    try:
        proc = subprocess.run([sys.executable, tmp], capture_output=True, text=True, timeout=timeout)
        if proc.returncode != 0:
            return False, (proc.stdout + "\n" + proc.stderr).strip()
        if not os.path.exists(output_path):
            return False, "output file was not created"
        return True, ""
    except subprocess.TimeoutExpired:
        return False, f"timeout after {timeout}s"
    finally:
        try: os.unlink(tmp)
        except OSError: pass


def find_test_cases(task_dir):
    cases = []
    for ip in sorted(glob.glob(os.path.join(task_dir, "*_input.xlsx"))):
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_input.xlsx", "_answer.xlsx")
        if os.path.exists(ap): cases.append((no, ip, ap))
    for ip in sorted(glob.glob(os.path.join(task_dir, "*_init.xlsx"))):
        no = os.path.basename(ip).split("_", 1)[0]
        ap = ip.replace("_init.xlsx", "_golden.xlsx")
        if os.path.exists(ap): cases.append((no, ip, ap))
    if not cases:
        bare_init = os.path.join(task_dir, "initial.xlsx")
        bare_gold = os.path.join(task_dir, "golden.xlsx")
        if os.path.exists(bare_init) and os.path.exists(bare_gold):
            cases.append(("1", bare_init, bare_gold))
    return cases


def load_items(path):
    if path.endswith(".json"):
        with open(path) as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = data.get("data") or list(data.values())
        return list(data)
    items = []
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line: items.append(json.loads(line))
    return items


# ── LLM call ────────────────────────────────────────────────────────────────

def llm_call(messages, deployment, max_tokens=16384, retries=5, llm_timeout=120):
    raw, _ = chat_messages_with_deployment(
        deployment=deployment,
        messages=messages,
        max_completion_tokens=max_tokens,
        retries=retries,
        stage="rollout",
        timeout=llm_timeout,
    )
    return str(raw or "")


# ── Process one task ────────────────────────────────────────────────────────

def process_one(item, data_root, out_root, model):
    task_id = str(item["id"])
    instruction = item["instruction"]
    instruction_type = item.get("instruction_type", "")
    answer_position = item.get("answer_position", "")
    answer_sheet = item.get("answer_sheet", "")
    if answer_position and answer_sheet and "!" not in answer_position:
        answer_position = f"{answer_sheet}!{answer_position}"

    sp = item.get("spreadsheet_path", f"spreadsheet/{task_id}")
    task_dir = sp if os.path.isabs(sp) else os.path.join(data_root, sp)

    result = {"id": task_id, "ok": False, "hard": 0, "soft": 0.0,
              "n_cases": 0, "n_pass": 0, "fail_reason": "", "error": ""}
    try:
        cases = find_test_cases(task_dir)
        result["n_cases"] = len(cases)
        if not cases:
            result["fail_reason"] = "no-test-cases"
            return result

        task_out = os.path.join(out_root, "predictions", task_id)
        os.makedirs(task_out, exist_ok=True)

        # LLM call
        system = build_system("")
        user = build_user(instruction, cases[0][1], instruction_type, answer_position)
        messages = [{"role": "system", "content": system}, {"role": "user", "content": user}]

        raw = llm_call(messages, model)
        time.sleep(3)
        code = extract_code(raw)

        with open(os.path.join(task_out, "code.py"), "w") as f: f.write(code)
        with open(os.path.join(task_out, "raw.txt"), "w") as f: f.write(raw)

        if not code.strip():
            result["fail_reason"] = "empty-code"
            return result

        # Execute + evaluate each test case
        for no, ip, ap in cases:
            pred = os.path.join(task_out, f"{no}_pred.xlsx")
            ok_exec, err = run_code(code, ip, pred)
            if not ok_exec:
                if not result["fail_reason"]:
                    result["fail_reason"] = f"exec: {err[:200]}"
                continue
            try:
                ev = evaluate(pred, ap, instruction_type, answer_position)
            except Exception as e:
                ev = {"ok": False, "reason": str(e)}
            if ev["ok"]:
                result["n_pass"] += 1

        nc, np = result["n_cases"], result["n_pass"]
        result["soft"] = np / nc if nc else 0.0
        result["hard"] = 1 if nc > 0 and np == nc else 0
        result["ok"] = bool(result["hard"])
        if result["ok"]: result["fail_reason"] = ""
        return result
    except Exception as e:
        result["fail_reason"] = f"unexpected: {e}"
        result["error"] = traceback.format_exc()
        return result


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Eval OFFICIAL prompt on verified-400")
    ap.add_argument("--model", default=MODEL)
    ap.add_argument("--backend", choices=["azure_openai", "codex", "claude"], default="azure_openai")
    ap.add_argument("--azure_endpoint", default="")
    ap.add_argument("--azure_api_version", default="")
    ap.add_argument("--azure_api_key", default="")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--out_root", default="")
    args = ap.parse_args()

    set_backend(args.backend)
    configure_azure_openai(
        endpoint=args.azure_endpoint or None,
        api_version=args.azure_api_version or None,
        api_key=args.azure_api_key or None,
    )
    set_student_deployment(args.model)
    ts = time.strftime("%Y%m%d_%H%M%S")
    out_root = args.out_root or os.path.join(_PROJECT_ROOT, "outputs", f"prompt_official_{args.model}_{ts}")
    out_root = os.path.abspath(out_root)
    os.makedirs(out_root, exist_ok=True)

    items = load_items(JSONL_PATH)
    if args.limit: items = items[:args.limit]

    print(f"{'='*60}")
    print(f"  Prompt: OFFICIAL (SpreadsheetBench original)")
    print(f"  Model:  {args.model}")
    print(f"  Items:  {len(items)}")
    print(f"  Output: {out_root}")
    print(f"{'='*60}")

    t0 = time.time()
    results = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(process_one, it, DATA_ROOT, out_root, args.model): it for it in items}
        for i, fut in enumerate(as_completed(futs), 1):
            item = futs[fut]
            try:
                res = fut.result(timeout=300)
            except FuturesTimeoutError:
                res = {"id": str(item["id"]), "ok": False, "hard": 0, "soft": 0.0,
                       "n_cases": 0, "n_pass": 0, "fail_reason": "timeout"}
            except Exception as e:
                res = {"id": str(item["id"]), "ok": False, "hard": 0, "soft": 0.0,
                       "n_cases": 0, "n_pass": 0, "fail_reason": str(e)}
            results.append(res)
            status = "PASS" if res.get("hard") else "FAIL"
            dt = time.time() - t0
            print(f"  {i}/{len(items)} id={res['id']:<10} {status}  cases={res.get('n_pass',0)}/{res.get('n_cases',0)}  dt={dt:.0f}s")

    # Summary
    hard_sum = sum(r.get("hard", 0) for r in results)
    soft_sum = sum(r.get("soft", 0.0) for r in results)
    n = len(results)
    print(f"\n{'='*60}")
    print(f"  OFFICIAL prompt: hard={hard_sum}/{n}={hard_sum/n:.4f}  soft={soft_sum/n:.4f}")
    print(f"{'='*60}")

    with open(os.path.join(out_root, "results.jsonl"), "w") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    with open(os.path.join(out_root, "summary.json"), "w") as f:
        json.dump({"prompt": "official", "model": args.model, "n": n,
                   "hard": hard_sum/n, "soft": soft_sum/n}, f, indent=2)


if __name__ == "__main__":
    main()
